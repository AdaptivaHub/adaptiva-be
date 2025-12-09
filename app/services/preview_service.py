import io
from typing import List, Dict, Any, Optional
from datetime import datetime, date, time
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
import pandas as pd
from fastapi import HTTPException

from app.utils import get_file_content
from app.models import PreviewResponse


def format_excel_cell_value(cell: Cell) -> str:
    """
    Format a cell value based on Excel's number format to match display.
    
    Args:
        cell: openpyxl Cell object
        
    Returns:
        Formatted string representation of the cell value
    """
    value = cell.value
    
    if value is None:
        return ""
    
    # Handle date/time types directly
    if isinstance(value, datetime):
        # Check the number format to determine display format
        fmt = cell.number_format or ""
        if "h" in fmt.lower() or "s" in fmt.lower():
            if "d" in fmt.lower() or "m" in fmt.lower() or "y" in fmt.lower():
                # Date and time
                return value.strftime("%m/%d/%Y %H:%M:%S")
            else:
                # Time only
                return value.strftime("%H:%M:%S")
        else:
            # Date only
            return value.strftime("%m/%d/%Y")
    
    if isinstance(value, date):
        return value.strftime("%m/%d/%Y")
    
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    
    # Handle numeric values with formatting
    if isinstance(value, (int, float)):
        fmt = cell.number_format or "General"
        
        # Handle percentage format
        if "%" in fmt:
            # Excel stores percentages as decimals (0.5 = 50%)
            percentage_value = value * 100
            # Count decimal places in format
            if "0.00%" in fmt:
                return f"{percentage_value:.2f}%"
            elif "0.0%" in fmt:
                return f"{percentage_value:.1f}%"
            elif "0%" in fmt:
                return f"{percentage_value:.0f}%"
            else:
                return f"{percentage_value:.2f}%"
        
        # Handle currency formats
        if "$" in fmt:
            if "#,##0.00" in fmt or "0.00" in fmt:
                return f"${value:,.2f}"
            elif "#,##0" in fmt:
                return f"${value:,.0f}"
            else:
                return f"${value:,.2f}"
        
        if "€" in fmt:
            return f"€{value:,.2f}"
        
        if "£" in fmt:
            return f"£{value:,.2f}"
        
        # Handle comma-separated numbers
        if "#,##0" in fmt:
            if ".00" in fmt:
                return f"{value:,.2f}"
            elif ".0" in fmt:
                return f"{value:,.1f}"
            else:
                return f"{value:,.0f}"
        
        # Handle decimal formats
        if "0.00" in fmt:
            return f"{value:.2f}"
        elif "0.0" in fmt:
            return f"{value:.1f}"
        elif fmt == "0":
            return f"{value:.0f}"
        
        # Handle scientific notation
        if "E" in fmt.upper():
            return f"{value:.2E}"
        
        # Default number formatting
        if isinstance(value, float):
            # Check if it's effectively an integer
            if value == int(value):
                return str(int(value))
            # Otherwise format with reasonable precision
            return f"{value:g}"
        
        return str(value)
    
    # Handle boolean
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    
    # Default: convert to string
    return str(value)


def get_excel_sheet_names(file_content: bytes) -> List[str]:
    """
    Get list of sheet names from an Excel file.
    
    Args:
        file_content: Raw bytes of the Excel file
        
    Returns:
        List of sheet names
    """
    workbook = load_workbook(io.BytesIO(file_content), read_only=True)
    sheet_names = workbook.sheetnames
    workbook.close()
    return sheet_names


def get_excel_preview_data(
    file_content: bytes,
    max_rows: int = 100,
    sheet_name: Optional[str] = None
) -> Dict[str, Any]:
    """
    Extract data from Excel file preserving display formatting.
    
    Args:
        file_content: Raw bytes of the Excel file
        max_rows: Maximum number of rows to return for preview
        sheet_name: Name of the sheet to preview (optional, defaults to first sheet)
        
    Returns:
                Dictionary with headers and formatted data rows
    """
    workbook = load_workbook(io.BytesIO(file_content), data_only=False)
    available_sheets = workbook.sheetnames
    
    # Select the appropriate sheet
    if sheet_name is not None:
        if sheet_name not in available_sheets:
            workbook.close()
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available sheets: {available_sheets}"
            )
        worksheet = workbook[sheet_name]
        active_sheet_name = sheet_name
    else:
        worksheet = workbook.active
        active_sheet_name = worksheet.title
    
    headers: List[str] = []
    data: List[Dict[str, str]] = []
    
    # Get the actual data range
    max_col = worksheet.max_column
    max_row = min(worksheet.max_row, max_rows + 1)  # +1 for header row
    
    # Extract headers from the first row
    for col in range(1, max_col + 1):
        cell = worksheet.cell(row=1, column=col)
        header_value = cell.value
        if header_value is not None:
            headers.append(str(header_value))
        else:
            headers.append(f"Column_{col}")
      # Extract data rows with formatting
    for row_idx in range(2, max_row + 1):
        row_data: Dict[str, str] = {}
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            row_data[header] = format_excel_cell_value(cell)
        data.append(row_data)
    
    workbook.close()
    
    return {
        "headers": headers,
        "data": data,
        "total_rows": worksheet.max_row - 1,  # Exclude header
        "formatted": True,
        "sheet_name": active_sheet_name,
        "available_sheets": available_sheets
    }


def get_csv_preview_data(
    file_content: bytes,
    max_rows: int = 100
) -> Dict[str, Any]:
    """
    Extract data from CSV file with basic formatting.
    
    Args:
        file_content: Raw bytes of the CSV file
        max_rows: Maximum number of rows to return for preview
        
    Returns:
        Dictionary with headers and data rows
    """    
    df = pd.read_csv(io.BytesIO(file_content), nrows=max_rows)
    
    # Get total row count by reading just the length
    total_df = pd.read_csv(io.BytesIO(file_content))
    total_rows = len(total_df)
    
    headers = df.columns.tolist()
    data = []
    
    for _, row in df.iterrows():
        row_data: Dict[str, str] = {}
        for header in headers:
            value = row[header]
            if pd.isna(value):
                row_data[header] = ""
            elif isinstance(value, float):
                # Check if it's effectively an integer
                if value == int(value):
                    row_data[header] = str(int(value))
                else:
                    row_data[header] = f"{value:g}"
            else:
                row_data[header] = str(value)
        data.append(row_data)
    
    return {
        "headers": headers,
        "data": data,
        "total_rows": total_rows,
        "formatted": False,
        "sheet_name": None,
        "available_sheets": None
    }


def get_preview_data(
    file_content: bytes,
    filename: str,
    max_rows: int = 100,
    sheet_name: Optional[str] = None
) -> Dict[str, Any]:
    """
    Get preview data from a file, preserving Excel formatting where possible.
    
    Args:
        file_content: Raw bytes of the file
        filename: Original filename to determine file type
        max_rows: Maximum number of rows to return
        sheet_name: Name of the sheet to preview (Excel only)
        
    Returns:
        Dictionary with headers, data, and metadata
    """
    filename_lower = filename.lower()
    
    if filename_lower.endswith('.xlsx') or filename_lower.endswith('.xls'):
        return get_excel_preview_data(file_content, max_rows, sheet_name)
    elif filename_lower.endswith('.csv'):
        return get_csv_preview_data(file_content, max_rows)
    else:
        raise ValueError(f"Unsupported file format: {filename}")


async def get_formatted_preview(file_id: str, max_rows: int = 100, sheet_name: Optional[str] = None):
    """
    Get formatted preview for an uploaded file.
    
    Args:
        file_id: The ID of the uploaded file
        max_rows: Maximum number of rows to return
        sheet_name: Name of the sheet to preview (Excel only)
        
    Returns:
        PreviewResponse with formatted data
    """
    # Get stored file content
    file_data = get_file_content(file_id)
    
    if file_data is None:
        raise HTTPException(
            status_code=404,
            detail=f"File with ID {file_id} not found"
        )
    
    content, filename = file_data
    
    try:
        preview_data = get_preview_data(content, filename, max_rows, sheet_name)
        
        return PreviewResponse(
            file_id=file_id,
            headers=preview_data["headers"],
            data=preview_data["data"],
            total_rows=preview_data["total_rows"],
            preview_rows=len(preview_data["data"]),
            formatted=preview_data["formatted"],
            message="Preview generated successfully",
            sheet_name=preview_data.get("sheet_name"),
            available_sheets=preview_data.get("available_sheets")
        )
    except ValueError as e:
        # Handle sheet not found errors
        raise HTTPException(
            status_code=400,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Error generating preview: {str(e)}"
        )
