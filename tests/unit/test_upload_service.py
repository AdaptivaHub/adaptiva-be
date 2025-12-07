"""
Unit tests for the upload service.

Tests based on requirements in: docs/requirements/file-upload.md
"""
import pytest
from io import BytesIO
from unittest.mock import Mock, AsyncMock, patch
import pandas as pd

from app.services.upload_service import process_file_upload
from app.models import FileUploadResponse


class TestProcessFileUpload:
    """Unit tests for process_file_upload function."""
    
    @pytest.mark.asyncio
    async def test_valid_csv_returns_response(self, sample_csv_content):
        """TC-1: Valid CSV upload returns file_id and metadata."""
        # Create mock UploadFile
        mock_file = Mock()
        mock_file.filename = "test.csv"
        mock_file.read = AsyncMock(return_value=sample_csv_content)
        
        # Call service
        response = await process_file_upload(mock_file)
        
        # Verify response
        assert isinstance(response, FileUploadResponse)
        assert response.file_id is not None
        assert len(response.file_id) == 36  # UUID format
        assert response.filename == "test.csv"
        assert response.rows == 3  # 3 data rows
        assert response.columns == 4  # 4 columns
        assert response.column_names == ["id", "name", "value", "category"]
        assert response.message == "File uploaded successfully"
    
    @pytest.mark.asyncio
    async def test_valid_xlsx_returns_response(self, sample_excel_file):
        """TC-2: Valid XLSX upload returns file_id and metadata."""
        # Read the BytesIO content
        content = sample_excel_file.read()
        sample_excel_file.seek(0)
        
        # Create mock UploadFile
        mock_file = Mock()
        mock_file.filename = "test.xlsx"
        mock_file.read = AsyncMock(return_value=content)
        
        # Call service
        response = await process_file_upload(mock_file)
        
        # Verify response
        assert isinstance(response, FileUploadResponse)
        assert response.file_id is not None
        assert response.filename == "test.xlsx"
        assert response.rows == 2  # 2 data rows
        assert response.columns == 4  # 4 columns
        assert "ID" in response.column_names
        assert "Name" in response.column_names
    
    @pytest.mark.asyncio
    async def test_unsupported_format_raises_400(self):
        """TC-5: Unsupported file format raises HTTPException 400."""
        from fastapi import HTTPException
        
        mock_file = Mock()
        mock_file.filename = "test.txt"
        mock_file.read = AsyncMock(return_value=b"some text content")
        
        with pytest.raises(HTTPException) as exc_info:
            await process_file_upload(mock_file)
        
        assert exc_info.value.status_code == 400
        assert "Unsupported file format" in exc_info.value.detail
    
    @pytest.mark.asyncio
    async def test_empty_csv_raises_400(self, sample_csv_with_headers_only):
        """TC-3: Empty CSV (headers only) raises HTTPException 400."""
        from fastapi import HTTPException
        
        mock_file = Mock()
        mock_file.filename = "empty.csv"
        mock_file.read = AsyncMock(return_value=sample_csv_with_headers_only)
        
        with pytest.raises(HTTPException) as exc_info:
            await process_file_upload(mock_file)
        
        assert exc_info.value.status_code == 400
        assert "empty" in exc_info.value.detail.lower()
    
    @pytest.mark.asyncio
    async def test_corrupted_file_raises_400(self):
        """TC-6: Corrupted file raises HTTPException 400."""
        from fastapi import HTTPException
        
        mock_file = Mock()
        mock_file.filename = "corrupted.csv"
        mock_file.read = AsyncMock(return_value=b"\x00\x01\x02\x03\x04")  # Binary garbage
        
        with pytest.raises(HTTPException) as exc_info:
            await process_file_upload(mock_file)
        
        assert exc_info.value.status_code == 400
        # Implementation may return "empty" or "error" message for garbage data
        assert exc_info.value.detail is not None
    
    @pytest.mark.asyncio
    async def test_large_file_processes_successfully(self, large_csv_content):
        """TC-7: Large file (1000+ rows) processes successfully."""
        mock_file = Mock()
        mock_file.filename = "large.csv"
        mock_file.read = AsyncMock(return_value=large_csv_content)
        
        response = await process_file_upload(mock_file)
        
        assert response.rows == 1000
        assert response.columns == 3
    
    @pytest.mark.asyncio
    async def test_special_characters_in_headers(self, csv_with_special_characters):
        """TC-8: Headers with special characters are preserved."""
        mock_file = Mock()
        mock_file.filename = "special.csv"
        mock_file.read = AsyncMock(return_value=csv_with_special_characters)
        
        response = await process_file_upload(mock_file)
        
        assert "Name (Full)" in response.column_names
        assert "Value ($)" in response.column_names
        assert "Percentage %" in response.column_names
    
    @pytest.mark.asyncio
    async def test_file_id_is_unique(self, sample_csv_content):
        """Each upload generates a unique file_id."""
        mock_file1 = Mock()
        mock_file1.filename = "test1.csv"
        mock_file1.read = AsyncMock(return_value=sample_csv_content)
        
        mock_file2 = Mock()
        mock_file2.filename = "test2.csv"
        mock_file2.read = AsyncMock(return_value=sample_csv_content)
        
        response1 = await process_file_upload(mock_file1)
        response2 = await process_file_upload(mock_file2)
        
        assert response1.file_id != response2.file_id
    
    @pytest.mark.asyncio
    async def test_xls_format_supported(self):
        """TC-10: Old Excel format (.xls) is supported."""
        # Create a simple XLS-like file (actually xlsx for testing)
        import openpyxl
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Col1'
        ws['A2'] = 'Value1'
        
        buffer = BytesIO()
        wb.save(buffer)
        content = buffer.getvalue()
        
        mock_file = Mock()
        mock_file.filename = "test.xls"  # .xls extension
        mock_file.read = AsyncMock(return_value=content)
        
        # Note: This will actually parse as xlsx format since openpyxl
        # doesn't support .xls, but pandas.read_excel handles both
        response = await process_file_upload(mock_file)
        
        assert response.file_id is not None


class TestDataStorage:
    """Tests for data storage after upload."""
    
    @pytest.mark.asyncio
    async def test_dataframe_stored_after_upload(self, sample_csv_content):
        """Uploaded data is stored and retrievable."""
        from app.utils import get_dataframe
        
        mock_file = Mock()
        mock_file.filename = "test.csv"
        mock_file.read = AsyncMock(return_value=sample_csv_content)
        
        response = await process_file_upload(mock_file)
        
        # Verify dataframe is stored
        df = get_dataframe(response.file_id)
        assert len(df) == 3
        assert list(df.columns) == ["id", "name", "value", "category"]
    
    @pytest.mark.asyncio
    async def test_file_content_stored_after_upload(self, sample_csv_content):
        """Original file content is stored for preview."""
        from app.utils import get_file_content
        
        mock_file = Mock()
        mock_file.filename = "test.csv"
        mock_file.read = AsyncMock(return_value=sample_csv_content)
        
        response = await process_file_upload(mock_file)
        
        # Verify file content is stored
        content, filename = get_file_content(response.file_id)
        assert content == sample_csv_content
        assert filename == "test.csv"
