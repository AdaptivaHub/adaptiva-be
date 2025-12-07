"""
Pytest configuration and shared fixtures for Adaptiva Backend tests.
"""
import pytest
from fastapi.testclient import TestClient
from io import BytesIO
import pandas as pd

# Import app
from app.main import app
from app.utils.storage import dataframes, file_contents


@pytest.fixture(autouse=True)
def clear_storage():
    """Clear in-memory storage before each test."""
    dataframes.clear()
    file_contents.clear()
    yield
    # Cleanup after test
    dataframes.clear()
    file_contents.clear()


@pytest.fixture
def client():
    """Create a test client for the FastAPI app."""
    return TestClient(app)


# =============================================================================
# CSV Fixtures
# =============================================================================

@pytest.fixture
def sample_csv_content() -> bytes:
    """Create sample CSV content for testing."""
    return b"id,name,value,category\n1,Item A,100.50,Electronics\n2,Item B,200.75,Clothing\n3,Item C,50.25,Electronics"


@pytest.fixture
def sample_csv_with_headers_only() -> bytes:
    """Create CSV with headers but no data rows."""
    return b"id,name,value,category"


@pytest.fixture
def large_csv_content() -> bytes:
    """Create a larger CSV for performance testing."""
    lines = ["id,name,value"]
    for i in range(1000):
        lines.append(f"{i},Item {i},{i * 10.5}")
    return "\n".join(lines).encode()


@pytest.fixture
def csv_with_special_characters() -> bytes:
    """Create CSV with special characters in headers and values."""
    return "Name (Full),Value ($),Percentage %\nJohn Doe,1234.56,15.5%\nJane Smith,9876.54,25.0%".encode()


# =============================================================================
# Excel Fixtures
# =============================================================================

@pytest.fixture
def sample_excel_file() -> BytesIO:
    """Create a sample Excel file with various data types."""
    import openpyxl
    from openpyxl.styles import numbers
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Headers
    ws['A1'] = 'ID'
    ws['B1'] = 'Name'
    ws['C1'] = 'Amount'
    ws['D1'] = 'Percentage'
    
    # Data row 1
    ws['A2'] = 1
    ws['B2'] = 'Product A'
    ws['C2'] = 1234.56
    ws['C2'].number_format = '"$"#,##0.00'
    ws['D2'] = 0.15
    ws['D2'].number_format = '0.00%'
    
    # Data row 2
    ws['A3'] = 2
    ws['B3'] = 'Product B'
    ws['C3'] = 9876.54
    ws['C3'].number_format = '"$"#,##0.00'
    ws['D3'] = 0.25
    ws['D3'].number_format = '0.00%'
    
    # Save to BytesIO
    file_buffer = BytesIO()
    wb.save(file_buffer)
    file_buffer.seek(0)
    
    return file_buffer


@pytest.fixture
def empty_excel_file() -> BytesIO:
    """Create an empty Excel file (no data)."""
    import openpyxl
    
    wb = openpyxl.Workbook()
    file_buffer = BytesIO()
    wb.save(file_buffer)
    file_buffer.seek(0)
    
    return file_buffer


@pytest.fixture
def excel_with_multiple_sheets() -> BytesIO:
    """Create Excel file with multiple sheets."""
    import openpyxl
    
    wb = openpyxl.Workbook()
    
    # First sheet (active)
    ws1 = wb.active
    ws1.title = "Sales"
    ws1['A1'] = 'Region'
    ws1['B1'] = 'Sales'
    ws1['A2'] = 'North'
    ws1['B2'] = 1000
    
    # Second sheet
    ws2 = wb.create_sheet("Costs")
    ws2['A1'] = 'Category'
    ws2['B1'] = 'Cost'
    ws2['A2'] = 'Labor'
    ws2['B2'] = 500
    
    file_buffer = BytesIO()
    wb.save(file_buffer)
    file_buffer.seek(0)
    
    return file_buffer


# =============================================================================
# Uploaded File Fixtures (Pre-uploaded, returns file_id)
# =============================================================================

@pytest.fixture
def uploaded_csv_file(client, sample_csv_content) -> str:
    """Upload a CSV file and return the file_id."""
    response = client.post(
        "/api/upload/",
        files={"file": ("test.csv", sample_csv_content, "text/csv")}
    )
    assert response.status_code == 200
    return response.json()["file_id"]


@pytest.fixture
def uploaded_excel_file(client, sample_excel_file) -> str:
    """Upload an Excel file and return the file_id."""
    response = client.post(
        "/api/upload/",
        files={"file": ("test.xlsx", sample_excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    )
    assert response.status_code == 200
    return response.json()["file_id"]


@pytest.fixture
def uploaded_large_csv(client, large_csv_content) -> str:
    """Upload a large CSV file and return the file_id."""
    response = client.post(
        "/api/upload/",
        files={"file": ("large.csv", large_csv_content, "text/csv")}
    )
    assert response.status_code == 200
    return response.json()["file_id"]


# =============================================================================
# DataFrame Fixtures (for unit tests)
# =============================================================================

@pytest.fixture
def sample_dataframe() -> pd.DataFrame:
    """Create a sample pandas DataFrame for testing."""
    return pd.DataFrame({
        'id': [1, 2, 3],
        'name': ['Item A', 'Item B', 'Item C'],
        'value': [100.50, 200.75, 50.25],
        'category': ['Electronics', 'Clothing', 'Electronics']
    })


@pytest.fixture
def numeric_dataframe() -> pd.DataFrame:
    """Create a DataFrame with numeric columns for chart testing."""
    return pd.DataFrame({
        'x': [1, 2, 3, 4, 5],
        'y': [10, 20, 15, 25, 30],
        'z': [5, 15, 10, 20, 25],
        'category': ['A', 'B', 'A', 'B', 'A']
    })
